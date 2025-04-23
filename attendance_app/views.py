from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from .models import AttendanceLog, DailyTimeAllocation
from django.utils import timezone
import csv
from datetime import date, datetime, timedelta, time
from django.db.models import Q
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.models import User
from collections import defaultdict
from django.utils.timezone import localtime
import pytz
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl.utils import get_column_letter

# 👤 LOGIN
def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_dashboard')  # Redirect superusers to admin dashboard
        else:
            return redirect('tracker')  # Redirect regular users to tracker

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            if user.is_superuser:
                return redirect('admin_dashboard')  # Redirect superusers to admin dashboard
            else:
                return redirect('tracker')  # Redirect regular users to tracker
        else:
            return render(request, 'login.html', {'error': 'Invalid credentials'})

    return render(request, 'login.html')

# 🚪 LOGOUT
def logout_view(request):
    logout(request)
    return redirect('login')

# 🧍 EMPLOYEE VIEW
@login_required
def tracker_view(request):
    # Check if a specific date was requested
    requested_date = request.GET.get('date')
    
    if requested_date:
        try:
            selected_date = datetime.strptime(requested_date, '%Y-%m-%d').date()
            today = selected_date
            is_today = (today == date.today())
        except ValueError:
            selected_date = today = date.today()
            is_today = True
    else:
        selected_date = today = date.today()
        is_today = True
    
    # Get or create daily time allocation for today
    time_allocation, created = DailyTimeAllocation.objects.get_or_create(
        user=request.user,
        date=today,
    )
    
    # Define shift hours (10 PM to 7 AM next day)
    now = localtime(timezone.now())
    current_date = now.date()
    
    # Create datetime objects for start and end of shift
    # If current time is before 7 AM, shift started yesterday at 10 PM
    if now.time() < time(7, 0):
        shift_start_date = current_date - timedelta(days=1)
        shift_end_date = current_date
    else:
        shift_start_date = current_date
        shift_end_date = current_date + timedelta(days=1)
        
    shift_start = datetime.combine(shift_start_date, time(22, 0)).replace(tzinfo=pytz.timezone('Asia/Manila'))
    shift_end = datetime.combine(shift_end_date, time(7, 0)).replace(tzinfo=pytz.timezone('Asia/Manila'))
    
    # Check for late clock in
    clocked_in_late = False
    minutes_late = 0
    first_clock_in = None

    # Retrieve logs for the user on the selected date
    logs = AttendanceLog.objects.filter(user=request.user, timestamp__date=today).order_by('-timestamp')
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(logs, 10)  # Show 10 activities per page
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Calculate current status
    current_status = 'idle'
    ongoing_break1 = False
    ongoing_break2 = False
    ongoing_lunch = False
    
    # Check if there's an ongoing break or lunch
    if time_allocation.break1_start_time:
        ongoing_break1 = True
        current_status = 'on_break1'
    
    if time_allocation.break2_start_time:
        ongoing_break2 = True
        current_status = 'on_break2'

    if time_allocation.lunch_start_time:
        ongoing_lunch = True
        current_status = 'at_lunch'

    # Check if user is clocked in across days
    # First, check if the user is clocked in today
    is_clocked_in = False
    clocked_in_today = AttendanceLog.objects.filter(
        user=request.user, 
        action='clock_in',
        timestamp__date=today
    ).exists()
    
    clocked_out_today = AttendanceLog.objects.filter(
        user=request.user, 
        action='clock_out',
        timestamp__date=today
    ).exists()
    
    # If viewing today, we also need to check if they're still clocked in from yesterday
    if is_today:
        # If not clocked in today, check if they clocked in yesterday but didn't clock out
        yesterday = date.today() - timedelta(days=1)
        if not clocked_in_today and not clocked_out_today:
            yesterday_clock_in = AttendanceLog.objects.filter(
                user=request.user,
                action='clock_in',
                timestamp__date=yesterday
            ).order_by('-timestamp').first()
            
            yesterday_clock_out = AttendanceLog.objects.filter(
                user=request.user,
                action='clock_out',
                timestamp__date=yesterday
            ).order_by('-timestamp').first()
            
            # If they clocked in yesterday and either didn't clock out or clocked in after clocking out
            if yesterday_clock_in and (not yesterday_clock_out or yesterday_clock_in.timestamp > yesterday_clock_out.timestamp):
                # They're still clocked in from yesterday
                is_clocked_in = True
                # Create a note that they're still clocked in from yesterday
                messages.info(request, f"You are still clocked in from yesterday ({yesterday.strftime('%Y-%m-%d')})")
    
    # Get the latest clock action for today
    latest_clock_action = AttendanceLog.objects.filter(
        user=request.user,
        action__in=['clock_in', 'clock_out'],
        timestamp__date=today
    ).order_by('-timestamp').first()

    # If the latest action is clock_in, they are clocked in
    if latest_clock_action and latest_clock_action.action == 'clock_in':
        is_clocked_in = True
        if not (ongoing_break1 or ongoing_break2 or ongoing_lunch):
            current_status = 'working'
    
    # Only allow clock actions on the current day, not in history
    can_perform_clock_actions = is_today

    # Handle POST request for action submission
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # CLOCK IN
        if action == 'clock_in':
            now = timezone.now()
            
            # Get Manila timezone
            manila_tz = pytz.timezone('Asia/Manila')
            now_manila = now.astimezone(manila_tz)
            
            # Calculate if late
            is_late = False
            late_minutes = 0
            
            # Determine shift date - if before 7 AM, shift is for previous day
            shift_date = now_manila.date()
            if now_manila.hour < 7:
                shift_date = shift_date - timedelta(days=1)
            
            # Calculate shift start time (10 PM on shift date)
            shift_start_time = time(22, 0)
            shift_start_datetime = datetime.combine(shift_date, shift_start_time)
            shift_start_datetime = pytz.timezone('Asia/Manila').localize(shift_start_datetime)
            
            # Check if clocked in after shift start time
            if now_manila > shift_start_datetime:
                is_late = True
                time_diff = now_manila - shift_start_datetime
                late_minutes = int(time_diff.total_seconds() // 60)
            
            # Create the log with late information if applicable
            note = None
            if is_late:
                note = f"Late arrival: {late_minutes} minutes"
            
            log = AttendanceLog(user=request.user, action='clock_in', note=note)
            log.save()
            
            if is_late:
                messages.warning(request, f'Clocked in {late_minutes} minutes late!')
            else:
                messages.success(request, 'Clocked in successfully!')
                
            return redirect('tracker')
        
        # CLOCK OUT
        elif action == 'clock_out':
            clock_out_note = request.POST.get('clock_out_note', '')
            log = AttendanceLog(user=request.user, action='clock_out', note=clock_out_note)
            log.save()
            messages.success(request, 'Clocked out successfully!')
            return redirect('tracker')
        
        # START BREAK 1
        elif action == 'start_break1':
            if not is_clocked_in:
                messages.error(request, 'You must be clocked in to take a break.')
            elif ongoing_break1 or ongoing_break2 or ongoing_lunch:
                messages.error(request, 'You are already on a break or lunch.')
            elif time_allocation.is_break1_exceeded():
                messages.error(request, 'You have already used all your break 1 time (15 minutes).')
            else:
                time_allocation.break1_start_time = timezone.now()
                time_allocation.save()
                log = AttendanceLog(user=request.user, action='start_break1')
                log.save()
                messages.success(request, 'Break 1 started!')
            return redirect('tracker')
        
        # START BREAK 2
        elif action == 'start_break2':
            if not is_clocked_in:
                messages.error(request, 'You must be clocked in to take a break.')
            elif ongoing_break1 or ongoing_break2 or ongoing_lunch:
                messages.error(request, 'You are already on a break or lunch.')
            elif time_allocation.is_break2_exceeded():
                messages.error(request, 'You have already used all your break 2 time (15 minutes).')
            else:
                time_allocation.break2_start_time = timezone.now()
                time_allocation.save()
                log = AttendanceLog(user=request.user, action='start_break2')
                log.save()
                messages.success(request, 'Break 2 started!')
            return redirect('tracker')
        
        # END BREAK 1
        elif action == 'end_break1':
            if not ongoing_break1:
                messages.error(request, 'No break 1 in progress.')
            else:
                # Calculate minutes used
                break_duration = timezone.now() - time_allocation.break1_start_time
                minutes_used = break_duration.total_seconds() // 60
                
                # Update time allocation
                time_allocation.break1_minutes_used += int(minutes_used)
                time_allocation.break1_start_time = None
                time_allocation.save()
                
                log = AttendanceLog(
                    user=request.user, 
                    action='end_break1',
                    note=f'Break 1 duration: {int(minutes_used)} minutes'
                )
                log.save()
                messages.success(request, f'Break 1 ended. You used {int(minutes_used)} minutes.')
            return redirect('tracker')
        
        # END BREAK 2
        elif action == 'end_break2':
            if not ongoing_break2:
                messages.error(request, 'No break 2 in progress.')
            else:
                # Calculate minutes used
                break_duration = timezone.now() - time_allocation.break2_start_time
                minutes_used = break_duration.total_seconds() // 60
                
                # Update time allocation
                time_allocation.break2_minutes_used += int(minutes_used)
                time_allocation.break2_start_time = None
                time_allocation.save()
                
                log = AttendanceLog(
                    user=request.user, 
                    action='end_break2',
                    note=f'Break 2 duration: {int(minutes_used)} minutes'
                )
                log.save()
                messages.success(request, f'Break 2 ended. You used {int(minutes_used)} minutes.')
            return redirect('tracker')
        
        # START LUNCH
        elif action == 'start_lunch':
            if not is_clocked_in:
                messages.error(request, 'You must be clocked in to take lunch.')
            elif ongoing_break1 or ongoing_break2 or ongoing_lunch:
                messages.error(request, 'You are already on a break or lunch.')
            elif time_allocation.is_lunch_exceeded():
                messages.error(request, 'You have already used all your lunch time (60 minutes).')
            else:
                time_allocation.lunch_start_time = timezone.now()
                time_allocation.save()
                log = AttendanceLog(user=request.user, action='start_lunch')
                log.save()
                messages.success(request, 'Lunch started!')
            return redirect('tracker')
        
        # END LUNCH
        elif action == 'end_lunch':
            if not ongoing_lunch:
                messages.error(request, 'No lunch in progress.')
            else:
                # Calculate minutes used
                lunch_duration = timezone.now() - time_allocation.lunch_start_time
                minutes_used = lunch_duration.total_seconds() // 60
                
                # Update time allocation
                time_allocation.lunch_minutes_used += int(minutes_used)
                time_allocation.lunch_start_time = None
                time_allocation.save()
                
                log = AttendanceLog(
                    user=request.user, 
                    action='end_lunch',
                    note=f'Lunch duration: {int(minutes_used)} minutes'
                )
                log.save()
                messages.success(request, f'Lunch ended. You used {int(minutes_used)} minutes.')
            return redirect('tracker')
    
    # Calculate percentages for progress bars
    break1_percentage = min(time_allocation.break1_minutes_used / 15 * 100, 100) if time_allocation.break1_minutes_used > 0 else 0
    break2_percentage = min(time_allocation.break2_minutes_used / 15 * 100, 100) if time_allocation.break2_minutes_used > 0 else 0
    lunch_percentage = min(time_allocation.lunch_minutes_used / 60 * 100, 100) if time_allocation.lunch_minutes_used > 0 else 0
    
    # Prepare context with all necessary data
    context = {
        'logs': logs,
        'page_obj': page_obj,
        'today': date.today(),
        'selected_date': selected_date,
        'is_today': is_today,
        'break1_minutes_remaining': time_allocation.break1_minutes_remaining(),
        'break2_minutes_remaining': time_allocation.break2_minutes_remaining(),
        'lunch_minutes_remaining': time_allocation.lunch_minutes_remaining(),
        'break1_minutes_exceeded': time_allocation.break1_minutes_exceeded(),
        'break2_minutes_exceeded': time_allocation.break2_minutes_exceeded(),
        'lunch_minutes_exceeded': time_allocation.lunch_minutes_exceeded(),
        'is_break1_exceeded': time_allocation.is_break1_exceeded(),
        'is_break2_exceeded': time_allocation.is_break2_exceeded(),
        'is_lunch_exceeded': time_allocation.is_lunch_exceeded(),
        'ongoing_break1': ongoing_break1,
        'ongoing_break2': ongoing_break2,
        'ongoing_lunch': ongoing_lunch,
        'current_status': current_status,
        'break1_minutes_used': time_allocation.break1_minutes_used,
        'break2_minutes_used': time_allocation.break2_minutes_used,
        'lunch_minutes_used': time_allocation.lunch_minutes_used,
        'break1_percentage': break1_percentage,
        'break2_percentage': break2_percentage,
        'lunch_percentage': lunch_percentage,
        'is_clocked_in': is_clocked_in,
        'can_perform_clock_actions': can_perform_clock_actions,
        'shift_start_time': '10:00 PM',
        'shift_end_time': '7:00 AM',
    }
    
    # Add timestamps for ongoing activities
    if ongoing_break1:
        context['break1_start_time'] = time_allocation.break1_start_time
    
    if ongoing_break2:
        context['break2_start_time'] = time_allocation.break2_start_time
    
    if ongoing_lunch:
        context['lunch_start_time'] = time_allocation.lunch_start_time
    
    return render(request, 'tracker.html', context)

# 📊 USER DASHBOARD
@login_required
def dashboard(request):
    logs = AttendanceLog.objects.filter(user=request.user).order_by('-timestamp')
    return render(request, 'dashboard.html', {'logs': logs})

# 📊 ADMIN DASHBOARD
@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    # Get filter parameters
    selected_user = request.GET.get('user')
    selected_date = request.GET.get('date')
    
    # Get page size from request (default to 5 if not specified)
    page_size = request.GET.get('size', '5')
    try:
        page_size = int(page_size)
        # Limit page size options to valid choices
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except ValueError:
        page_size = 5
    
    # Base query for logs
    logs_query = AttendanceLog.objects.select_related('user')
    
    # Apply filters
    if selected_user:
        logs_query = logs_query.filter(user_id=selected_user)
        selected_user_obj = User.objects.get(id=selected_user)
    else:
        selected_user_obj = None
        
    if selected_date:
        try:
            filter_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            logs_query = logs_query.filter(timestamp__date=filter_date)
        except ValueError:
            pass
    
    # Order logs
    logs = logs_query.order_by('-timestamp')
    
    # Pagination with the selected page size
    page = request.GET.get('page', 1)
    paginator = Paginator(logs, page_size)
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Get all users (for filter dropdown and sidebar)
    users = User.objects.all().order_by('username')
    
    # Calculate current stats
    today = date.today()
    total_users = User.objects.count()
    
    # Track user statuses for sidebar
    clocked_in_users = set()
    break_users = set()
    lunch_users = set()
    
    # Find users currently clocked in, on break, or lunch
    for user_obj in users:
        # Check if user is clocked in
        clock_ins_today = AttendanceLog.objects.filter(
            user=user_obj,
            action='clock_in',
            timestamp__date=today
        ).count()
        
        clock_outs_today = AttendanceLog.objects.filter(
            user=user_obj,
            action='clock_out',
            timestamp__date=today
        ).count()
        
        if clock_ins_today > clock_outs_today:
            clocked_in_users.add(user_obj.id)
            
            # Check if they're on break or lunch
            try:
                time_allocation = DailyTimeAllocation.objects.get(user=user_obj, date=today)
                if time_allocation.break1_start_time or time_allocation.break2_start_time:
                    break_users.add(user_obj.id)
                if time_allocation.lunch_start_time:
                    lunch_users.add(user_obj.id)
            except DailyTimeAllocation.DoesNotExist:
                pass
    
    # Get time allocations for all users for their log dates
    # Create a nested dictionary: {user_id: {date: allocation}}
    user_time_allocations = defaultdict(dict)
    
    # Get unique user and date combinations from the logs
    user_date_pairs = logs.values('user_id', 'timestamp__date').distinct()
    
    for pair in user_date_pairs:
        user_id = pair['user_id']
        log_date = pair['timestamp__date']
        
        # Get or create time allocation for this user and date
        time_allocation, _ = DailyTimeAllocation.objects.get_or_create(
            user_id=user_id,
            date=log_date
        )
        
        # Store in our dictionary
        user_time_allocations[user_id][log_date] = time_allocation
    
    context = {
        'logs': page_obj,
        'page_obj': page_obj,
        'users': users,
        'selected_user': selected_user,
        'selected_user_obj': selected_user_obj,
        'selected_date': selected_date,
        'total_users': total_users,
        'users_clocked_in': len(clocked_in_users),
        'users_on_break': len(break_users),
        'users_on_lunch': len(lunch_users),
        'user_time_allocations': dict(user_time_allocations),
        'clocked_in_users': clocked_in_users,
        'break_users': break_users,
        'lunch_users': lunch_users,
    }
    
    return render(request, 'admin_dashboard.html', context)

# 📤 EXPORT USER CSV
@login_required
def export_csv(request):
    # Import the necessary libraries
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    # Create a new workbook and sheets
    wb = openpyxl.Workbook()
    
    # Create one main activity sheet
    activity_sheet = wb.active
    activity_sheet.title = "Activity Log"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set up headers for activity sheet
    headers = ["Date", "Time", "Action", "Status", "Note"]
    
    # Apply headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = activity_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get all logs for the user, ordered by timestamp
    logs = AttendanceLog.objects.filter(user=request.user).order_by('-timestamp')
    
    # Process each log entry
    row = 2
    for log in logs:
        # Format date and time
        log_date = timezone.localtime(log.timestamp).strftime('%Y-%m-%d')
        log_time = timezone.localtime(log.timestamp).strftime('%I:%M %p')
        
        # Format the action nicely
        action = log.action.replace('_', ' ').title()
        
        # Determine status based on action
        if 'Clock In' in action:
            status = "Working"
        elif 'Clock Out' in action:
            status = "Off Duty"
        elif 'Start Break' in action or 'Start Lunch' in action:
            status = "On Break"
        elif 'End Break' in action or 'End Lunch' in action:
            status = "Back to Work"
        else:
            status = ""
        
        # Apply alternating row styling
        row_style = alt_row_fill if row % 2 == 0 else None
        
        # Apply styling to all cells in the row
        for col in range(1, 6):
            cell = activity_sheet.cell(row=row, column=col)
            cell.border = thin_border
            if row_style:
                cell.fill = row_style
        
        # Add data to the sheet
        activity_sheet.cell(row=row, column=1).value = log_date
        activity_sheet.cell(row=row, column=2).value = log_time
        activity_sheet.cell(row=row, column=3).value = action
        activity_sheet.cell(row=row, column=4).value = status
        activity_sheet.cell(row=row, column=5).value = log.note or ""
        
        row += 1
    
    # Adjust column widths using helper function
    adjust_column_widths(activity_sheet)
    
    # Set up print settings
    activity_sheet.page_setup.orientation = activity_sheet.ORIENTATION_LANDSCAPE
    activity_sheet.page_setup.fitToWidth = True
    activity_sheet.print_title_rows = '1:1'  # Repeat header row on each page
    
    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Reset the pointer to the beginning

    # Return response with the current month and year in the filename
    current_month_year = datetime.now().strftime('%B_%Y')
    filename = f"{request.user.username}_activity_log_{current_month_year}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

# 📤 EXPORT ADMIN CSV (ALL USERS)
@user_passes_test(lambda u: u.is_superuser)
def admin_export_csv(request):
    # Import necessary libraries
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    # Get specific user parameter if provided
    selected_user_id = request.GET.get('user')
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet that's created automatically
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create the summary sheet first
    summary_sheet = wb.create_sheet(title="Summary")
    
    # Define styles
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set up headers for summary sheet - simplified to match user export
    summary_headers = ["Username", "Date", "Time", "Action", "Status", "Note"]
    
    # Apply headers to summary sheet with styling
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get data depending on whether we're showing all users or a specific user
    if selected_user_id:
        users = User.objects.filter(id=selected_user_id)
        filename_prefix = f"user_{users.first().username}_activity"
    else:
        users = User.objects.all().order_by('username')
        filename_prefix = "all_users_activity"
    
    # Keep track of created sheet names to avoid duplicates
    created_sheet_names = set(["Summary"])
    
    # Track the current row in the summary sheet
    summary_row = 2
    
    # Process each user's data
    for user in users:
        # Get all logs for the user, ordered by timestamp
        logs = AttendanceLog.objects.filter(user=user).order_by('-timestamp')
        
        # If user has no logs, continue to next user
        if not logs.exists():
            continue
        
        # Create a unique name for this user's sheet
        base_sheet_name = f"{user.username[:20]}"  # Limit to 20 chars for Excel sheet name limit
        sheet_name = base_sheet_name
        counter = 1
        
        # Ensure unique sheet names by adding numbers if needed
        while sheet_name in created_sheet_names:
            sheet_name = f"{base_sheet_name}_{counter}"
            counter += 1
        
        # Add to our set of created sheet names
        created_sheet_names.add(sheet_name)
        
        # Create a separate sheet for this user
        user_sheet = wb.create_sheet(title=sheet_name)
        
        # Set up headers for user sheet - matched to user export
        user_headers = ["Date", "Time", "Action", "Status", "Note"]
        
        # Add header with user information
        user_sheet.merge_cells('A1:E1')
        user_header = user_sheet.cell(row=1, column=1)
        user_header.value = f"Activity Log for: {user.first_name} {user.last_name} ({user.username})"
        user_header.font = Font(bold=True, size=14)
        user_header.alignment = Alignment(horizontal='center')
        
        # Apply headers with styling (in row 2)
        for col_num, header in enumerate(user_headers, 1):
            cell = user_sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Process each log entry for both the user sheet and summary sheet
        user_row = 3  # Start from row 3 since row 1-2 have headers
        
        for log in logs:
            # Format date and time
            log_date = timezone.localtime(log.timestamp).strftime('%Y-%m-%d')
            log_time = timezone.localtime(log.timestamp).strftime('%I:%M %p')
            
            # Format the action nicely
            action = log.action.replace('_', ' ').title()
            
            # Determine status based on action - same as user export
            if 'Clock In' in action:
                status = "Working"
            elif 'Clock Out' in action:
                status = "Off Duty"
            elif 'Start Break' in action or 'Start Lunch' in action:
                status = "On Break"
            elif 'End Break' in action or 'End Lunch' in action:
                status = "Back to Work"
            else:
                status = ""
            
            # Apply alternating row styling for user sheet
            row_style = alt_row_fill if user_row % 2 == 0 else None
            
            # Apply styling to all cells in the row for user sheet
            for col in range(1, 6):
                cell = user_sheet.cell(row=user_row, column=col)
                cell.border = thin_border
                if row_style:
                    cell.fill = row_style
            
            # Add data to the user sheet
            user_sheet.cell(row=user_row, column=1).value = log_date
            user_sheet.cell(row=user_row, column=2).value = log_time
            user_sheet.cell(row=user_row, column=3).value = action
            user_sheet.cell(row=user_row, column=4).value = status
            user_sheet.cell(row=user_row, column=5).value = log.note or ""
            
            user_row += 1
            
            # Now add the same data to the summary sheet
            # Apply alternating row styling for summary sheet
            row_style = alt_row_fill if summary_row % 2 == 0 else None
            
            # Apply styling to all cells in the row for summary sheet
            for col in range(1, 7):
                cell = summary_sheet.cell(row=summary_row, column=col)
                cell.border = thin_border
                if row_style:
                    cell.fill = row_style
            
            # Add data to the summary sheet
            summary_sheet.cell(row=summary_row, column=1).value = user.username
            summary_sheet.cell(row=summary_row, column=2).value = log_date
            summary_sheet.cell(row=summary_row, column=3).value = log_time
            summary_sheet.cell(row=summary_row, column=4).value = action
            summary_sheet.cell(row=summary_row, column=5).value = status
            summary_sheet.cell(row=summary_row, column=6).value = log.note or ""
            
            summary_row += 1
        
        # Use our helper function to adjust column widths safely for user sheet
        adjust_column_widths(user_sheet)
        
        # Set print area and page setup for user sheet
        user_sheet.page_setup.orientation = user_sheet.ORIENTATION_LANDSCAPE
        user_sheet.page_setup.fitToWidth = True
        user_sheet.print_title_rows = '1:2'  # Repeat header rows on each page
    
    # Use our helper function to adjust column widths safely for summary sheet
    adjust_column_widths(summary_sheet)
    
    # Set summary sheet as the active sheet
    wb.active = summary_sheet
    
    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return response with appropriate filename
    current_month_year = datetime.now().strftime('%B_%Y')
    filename = f"{filename_prefix}_{current_month_year}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# Modified function to safely handle column width adjustment

def adjust_column_widths(sheet):
    """Helper function to safely adjust column widths"""
    for col_idx in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Check each cell in the column
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            # Only process if it's a normal cell (not merged) and has a value
            if cell.__class__.__name__ != 'MergedCell' and cell.value:
                try:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
                except:
                    pass
        
        # Set column width
        adjusted_width = max(max_length + 2, 12)  # Minimum width of 12
        sheet.column_dimensions[column_letter].width = adjusted_width

# Automatic clock out function
def auto_clock_out_at_shift_end():
    """
    Automatically clock out users at 7 AM if they're still clocked in,
    end any ongoing breaks or lunch, and reset break/lunch allocations for the new day
    """
    now = timezone.now()
    manila_tz = pytz.timezone('Asia/Manila')
    now_manila = now.astimezone(manila_tz)
    
    # Rest of the function...
